from openpyxl import load_workbook
import re
import traceback


class product_formatter:
    format_list=[('Модель',''),('"',''),('matte',''),('violet',''),('Chocolate','')]

    def format_product(self,product):
        for i in self.format_list:
            product=product.replace(i[0],i[1])
        # print(product)
        return product

class Xlsx_parser:
    ucenka_cols={'name':0,'opt_price':1}
    
    
    def get_ucenka_table(self,file_path):
       
        wb = load_workbook((file_path))
        sheet=wb.sheetnames[0]      
        table=wb[sheet]
        data={}
        product_names=[]
       
        for i in table.rows:
            # print(i)
            
            try:
                try:
                    opt_price=(i[self.ucenka_cols['opt_price']].value)
                    # print(opt_price)
                    try:
                        # print(opt_price)
                        opt_price=int( opt_price)
                    except:
                        opt_price=int(re.sub('\D', '', opt_price))
                    name=(i[self.ucenka_cols['name']].value)
                    assert opt_price != None
    
                except Exception as e:
                    traceback.print_exc()

                    # print(e,'IN EXCEL PARSER')
                    raise Exception(e)
                name=(product_formatter().format_product(name))
                data[name]={
                    
                        'name':(name),
                        'opt_price':opt_price}
            except Exception as e:
                print('213',e)
                continue
            
        
        return data
if __name__=='__main__':
    print(Xlsx_parser().get_ucenka_table('margin.xlsx'))
    # print(product_formatter().format_product('Модель "VCB 0316" blue'))